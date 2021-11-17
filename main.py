import openpyxl
import requests


# lists out which links have denied access in excel file
def filter_links(file):
    wbkName = file + '.xlsx'
    wrkbk = openpyxl.load_workbook(wbkName)
    sheet = wrkbk.get_sheet_by_name(file)
    for i in range(2, sheet.max_row + 1):
        cell_obj = sheet.cell(row=i, column=1)
        url = str(cell_obj.value)
        req = requests.get(url, 'html.parser')
        if (req.text[40:45] == 'Error'):
            coords = "C" + str(i)
            sheet[coords] = "Access Denied"

    wrkbk.save(wbkName)
    wrkbk.close


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    fileName = input("input excel file name (make sure its the same as the sheet name): ")
    filter_links(fileName)
